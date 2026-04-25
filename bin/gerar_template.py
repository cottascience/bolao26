#!/usr/bin/env python3
"""Gera assets/palpites_template.xlsx a partir de _data/matches.csv.

Layout (por linha):
    | # | Quando | Mandante | gols_m | × | gols_v | Visitante |

72 partidas da fase de grupos, agrupadas A→L com cabeçalho de grupo.
Células de input (gols) destacadas; resto somente leitura por convenção.
Validação de dados: inteiro 0–20.
"""

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "_data"
OUT = ROOT / "assets" / "palpites_template.xlsx"

COL_WIDTHS = {"A": 5, "B": 14, "C": 28, "D": 8, "E": 4, "F": 8, "G": 28}
HEADER_BG = "2D3250"
GROUP_BG = "F4E5DC"
SCORE_BG = "FFF4D6"
TEXT_DARK = "2D3250"
TEXT_LIGHT = "FFF9F5"
TEXT_MUTED = "7A7F9C"


def load(path):
    with open(path, newline="") as f:
        return list(csv.DictReader(f))


def main():
    matches = load(DATA / "matches.csv")
    teams = {t["id"]: t for t in load(DATA / "teams.csv")}

    grupo = [m for m in matches if m["stage_id"] == "1"]
    assert len(grupo) == 72, f"esperava 72, achei {len(grupo)}"

    by_group = {}
    for m in grupo:
        g = teams[m["home_team_id"]]["group_letter"]
        by_group.setdefault(g, []).append(m)
    for g in by_group:
        by_group[g].sort(key=lambda m: int(m["match_number"]))

    wb = Workbook()
    ws = wb.active
    ws.title = "Palpites"

    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    header_font = Font(name="Calibri", size=11, bold=True, color=TEXT_LIGHT)
    header_fill = PatternFill("solid", fgColor=HEADER_BG)

    name_label = ws.cell(row=1, column=1, value="NOME")
    name_label.font = header_font
    name_label.fill = header_fill
    name_label.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=7)
    name_input = ws.cell(row=1, column=2)
    name_input.fill = PatternFill("solid", fgColor=SCORE_BG)
    name_input.font = Font(name="Calibri", size=14, bold=True, color=TEXT_DARK)
    name_input.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30

    for i, h in enumerate(["#", "Quando", "Mandante", "Gols", "×", "Gols", "Visitante"], 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 24
    ws.freeze_panes = "A3"

    group_font = Font(name="Calibri", size=15, bold=True, color=TEXT_DARK)
    group_fill = PatternFill("solid", fgColor=GROUP_BG)
    team_font = Font(name="Calibri", size=14, bold=True, color=TEXT_DARK)
    score_font = Font(name="Calibri", size=14, bold=True, color=TEXT_DARK)
    score_fill = PatternFill("solid", fgColor=SCORE_BG)
    meta_font = Font(name="Calibri", size=11, color=TEXT_MUTED)
    x_font = Font(name="Calibri", size=14, bold=True, color=TEXT_MUTED)

    dv = DataValidation(
        type="whole", operator="between", formula1=0, formula2=20,
        showErrorMessage=True,
        errorTitle="Gol inválido",
        error="Use um número inteiro de 0 a 20.",
    )
    ws.add_data_validation(dv)

    row = 3
    for g in sorted(by_group):
        ws.cell(row=row, column=1, value=f"Grupo {g}")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        gc = ws.cell(row=row, column=1)
        gc.font = group_font
        gc.fill = group_fill
        gc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 28
        row += 1

        for m in by_group[g]:
            home = teams[m["home_team_id"]]["team_name_pt"]
            away = teams[m["away_team_id"]]["team_name_pt"]
            date_part, time_part = m["kickoff_brt"].split(" ", 1)
            _, mo, d = date_part.split("-")
            hh, mm, _ = time_part.split(":", 2)
            quando = f"{d}/{mo} {hh}:{mm}"

            ws.cell(row=row, column=1, value=int(m["id"])).font = meta_font
            ws.cell(row=row, column=2, value=quando).font = meta_font
            ws.cell(row=row, column=3, value=home).font = team_font
            ws.cell(row=row, column=5, value="×").font = x_font
            ws.cell(row=row, column=7, value=away).font = team_font

            aligns = {1: "center", 2: "center", 3: "right", 5: "center", 7: "left"}
            for col, h in aligns.items():
                ws.cell(row=row, column=col).alignment = Alignment(
                    horizontal=h, vertical="center"
                )

            for col in (4, 6):
                cell = ws.cell(row=row, column=col)
                cell.fill = score_fill
                cell.font = score_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                dv.add(cell)

            ws.row_dimensions[row].height = 24
            row += 1

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {len(grupo)} matches, {len(by_group)} groups -> {OUT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
