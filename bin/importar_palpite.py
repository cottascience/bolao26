#!/usr/bin/env python3
"""Valida um palpite preenchido (xlsx) e escreve _data/palpites/<slug>.yml.

Uso:
    bin/importar_palpite.py caminho/do/arquivo.xlsx [--apelido <override>]

- Nome do participante vem do campo "NOME" na primeira linha do arquivo.
- Slug gerado a partir do nome (acentos removidos, espaços viram '-').
- Schema: 72 linhas (match_id 1..72) com gols inteiros >= 0.
- Saída: YAML com `nome`, `slug`, `recebido_em`, e lista de `palpites`.
"""

import argparse
import csv
import datetime as dt
import re
import sys
import unicodedata
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "_data"
OUT_DIR = DATA / "palpites"


def fail(msg):
    print(f"ERRO: {msg}", file=sys.stderr)
    sys.exit(1)


def slugify(s):
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s or "anonimo"


def parse_xlsx(path, matches_by_id):
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb.active

    nome = ws.cell(row=1, column=2).value
    if not nome or not isinstance(nome, str) or not nome.strip():
        fail("preencha o campo 'NOME' no topo do arquivo")
    nome = " ".join(nome.split())

    palpites = []
    for r in range(3, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if not isinstance(v, int):
            continue
        mid = v
        if mid not in matches_by_id:
            fail(f"linha {r}: match_id {mid} não existe na fase de grupos")
        gm = ws.cell(row=r, column=4).value
        gv = ws.cell(row=r, column=6).value
        for val, name in [(gm, "mandante"), (gv, "visitante")]:
            if val is None or (isinstance(val, str) and val.strip() == ""):
                fail(f"linha {r} (jogo {mid}): gols do {name} vazio")
            try:
                n = int(val)
            except (ValueError, TypeError):
                fail(f"linha {r} (jogo {mid}): gols do {name} = {val!r} não é inteiro")
            if n < 0:
                fail(f"linha {r} (jogo {mid}): gols do {name} = {n} é negativo")
        m = matches_by_id[mid]
        palpites.append({
            "id": mid,
            "grupo": m["grupo"],
            "data": m["data"],
            "hora": m["hora"],
            "mandante": m["mandante"],
            "visitante": m["visitante"],
            "gm": int(gm),
            "gv": int(gv),
        })
    return nome, palpites


def write_yaml(dest, nome, slug, palpites):
    with open(dest, "w", encoding="utf-8") as f:
        f.write(f"nome: {nome}\n")
        f.write(f"slug: {slug}\n")
        f.write(f"recebido_em: {dt.date.today().isoformat()}\n")
        f.write("palpites:\n")
        for p in palpites:
            f.write(
                f"  - {{ id: {p['id']:>2}, grupo: {p['grupo']}, data: {p['data']}, hora: \"{p['hora']}\", "
                f"mandante: {yaml_str(p['mandante'])}, visitante: {yaml_str(p['visitante'])}, "
                f"gm: {p['gm']}, gv: {p['gv']} }}\n"
            )


def yaml_str(s):
    if any(c in s for c in ":#&*!|>'\"%@`,[]{}") or s.strip() != s:
        return '"' + s.replace("\\", "\\\\").replace('"', '\\"') + '"'
    return s


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("arquivo")
    ap.add_argument("--apelido", help="override do slug derivado do nome")
    args = ap.parse_args()

    src = args.arquivo
    if Path(src).suffix.lower() != ".xlsx":
        fail(f"esperava .xlsx, recebi '{Path(src).suffix}'")

    with open(DATA / "matches.csv", newline="") as f:
        matches_rows = list(csv.DictReader(f))
    with open(DATA / "teams.csv", newline="") as f:
        teams = {t["id"]: t for t in csv.DictReader(f)}

    matches_by_id = {}
    for m in matches_rows:
        if m["stage_id"] != "1":
            continue
        date_part, time_part = m["kickoff_brt"].split(" ", 1)
        _, mo, d = date_part.split("-")
        hh, mm, _ = time_part.split(":", 2)
        matches_by_id[int(m["id"])] = {
            "data": f"{d}/{mo}",
            "hora": f"{hh}:{mm}",
            "mandante": teams[m["home_team_id"]]["team_name_pt"],
            "visitante": teams[m["away_team_id"]]["team_name_pt"],
            "grupo": teams[m["home_team_id"]]["group_letter"],
        }

    nome, palpites = parse_xlsx(src, matches_by_id)

    if len(palpites) != 72:
        fail(f"esperava 72 palpites, achei {len(palpites)}")
    seen = {p["id"] for p in palpites}
    missing = set(matches_by_id) - seen
    if missing:
        fail(f"faltam match_ids: {sorted(missing)}")
    palpites.sort(key=lambda p: p["id"])

    slug = args.apelido or slugify(nome)
    if not slug.replace("-", "").replace("_", "").isalnum():
        fail(f"slug '{slug}' inválido (só alfanumérico, '-', '_')")

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    dest = OUT_DIR / f"{slug}.yml"
    if dest.exists():
        print(f"AVISO: sobrescrevendo {dest.relative_to(ROOT)}")
    write_yaml(dest, nome, slug, palpites)
    print(f"OK: {nome} ({slug}) — {len(palpites)} palpites -> {dest.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
